import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def adjust_column_widths(ws):
    """
    Adjust the width of each column in the worksheet to fit the content.
    Iterates through all cells in each column to find the maximum content length
    and sets the column width accordingly.
    """
    column_widths = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                column = cell.column_letter
                try:
                    cell_length = len(str(cell.value))
                except:
                    cell_length = 0
                if column in column_widths:
                    column_widths[column] = max(column_widths[column], cell_length)
                else:
                    column_widths[column] = cell_length
    
    for column, width in column_widths.items():
        adjusted_width = max(width + 2, 10)
        ws.column_dimensions[column].width = adjusted_width

def main():
    # Get the current date for the output filename
    current_date = datetime.now().strftime("%m-%d-%y")

    # Define the paths to the Excel files with their full paths
    files = [ 
        f"_results/Informe_Semana_Epidemiologica_DATA/Informe_Semana_Epidemiologica_{current_date}.xlsx",
        f"_results/Dengue_Cases_Year_DATA/Dengue_Cases_Year_{current_date}.xlsx",
        f"_results/Big_Numbers_DATA/Big_Numbers_UF_{current_date}.xlsx",
        f"_results/SE_COMPLETA_2023-24_DATA/SE_COMPLETA_2023-24_{current_date}.xlsx",       
        f"_results/City_Cases_2024_DATA/City_Cases_{current_date}.xlsx"    
    ]

    # Define the fixed sheet names corresponding to each file
    fixed_sheet_names = [
        "Informe Semana Epidemiologica",
        "Dengue Cases Year",
        "Big Numbers UF",
        "SE Completa 2023-24",     
        "City Cases 2024"
    ]

    # Create a new workbook for the merged file
    wb = Workbook()
    # Remove the default sheet created by Workbook()
    wb.remove(wb.active)

    # Copy sheets from each source file into the new workbook
    for file_path, sheet_name in zip(files, fixed_sheet_names):
        if not os.path.exists(file_path):
            print(f"Error: File not found: {file_path}")
            continue

        # Load the source workbook
        source_wb = load_workbook(file_path)
        
        # Assuming each source file has only one sheet, get the first sheet
        source_ws = source_wb.active
        
        # Create a new sheet in the destination workbook with the fixed sheet name
        dest_ws = wb.create_sheet(title=sheet_name)

        # Copy all cells from the source sheet to the destination sheet
        for row in source_ws.rows:
            for cell in row:
                new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()

        # Copy column dimensions
        for col in source_ws.column_dimensions:
            dest_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

        # Copy row dimensions
        for row in source_ws.row_dimensions:
            dest_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

        # Adjust column widths to fit content (as a fallback if original widths aren't set)
        adjust_column_widths(dest_ws)

    # Save the merged workbook
    output_path = f"_results/_final_output/Epidemiology_Dengue_{current_date}.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)  # Ensure output directory exists
    wb.save(output_path)
    print(f"Merged file saved as: {output_path}")

if __name__ == "__main__":
    main()