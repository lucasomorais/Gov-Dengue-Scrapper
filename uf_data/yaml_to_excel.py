import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

def yaml_to_excel_with_exact_formatting(yaml_file, excel_file):
    """
    Converts a YAML file containing UF data into an Excel file with the exact format
    matching Pasta1.xlsx (blue headers, alternating row colors, centered data, gridlines).
    """
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    # Define styles for headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define border style for gridlines
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # Add headers
    headers = ["Casos 2025", "Incidence 2025", "Óbitos 2025 (em investigação)", "Óbitos 2025 (confirmados)"]
    ws.append(headers)

    # Format header row
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    ws.row_dimensions[1].height = 22  # Match header height from Pasta1.xlsx

    # Load YAML data
    with open(yaml_file, "r", encoding="utf-8") as f:
        uf_data = yaml.safe_load(f)

    # Define alternating row colors
    light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Light blue
    light_orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light orange
    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows with alternating colors
    for row_idx, (uf, data) in enumerate(uf_data.items(), 2):
        casos = data.get("Casos prováveis de Dengue", "").replace(",", "")
        incidencia = data.get("Coeficiente de incidência", "").replace(",", ".")
        obitos_investigacao = data.get("Óbitos em investigação", "")
        obitos_confirmados = data.get("Óbitos por Dengue", "")

        row = [
            int(casos) if casos else 0,
            float(incidencia) if incidencia else 0.0,
            int(obitos_investigacao) if obitos_investigacao else 0,
            int(obitos_confirmados) if obitos_confirmados else 0
        ]
        ws.append(row)

        # Apply alternating row colors
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            if col_idx == 2:  # Incidence column
                cell.number_format = "#,##0.0"  # One decimal place
            else:
                cell.number_format = "#,##0"  # Integer with thousand separator
            # Apply fill based on row index (odd/even)
            if row_idx % 2 == 0:  # Even rows (light blue)
                cell.fill = light_blue_fill
            else:  # Odd rows (light orange)
                cell.fill = light_orange_fill
        ws.row_dimensions[row_idx].height = 16  # Match data row height from Pasta1.xlsx

    # Adjust column widths dynamically
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 15)  # Cap width at 15 for consistency
        ws.column_dimensions[column].width = adjusted_width

    # Enable gridlines
    ws.sheet_view.showGridLines = True

    # Save the workbook
    wb.save(excel_file)
    print(f"Excel file saved to {excel_file}")

# Example usage
yaml_to_excel_with_exact_formatting("dengue_uf_data.yaml", "epidemiology_uf_data.xlsx")