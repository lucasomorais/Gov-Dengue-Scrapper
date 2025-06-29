import pytest, os, sys, yaml, shutil, re, glob
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from modules.utils import (
    find_latest_file_in_subfolders,
    OUTPUT_DIR,
    TEMP_DIR,
)


TEMP_DIR = os.path.join(BASE_DIR, 'temp')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
DENGUE_DIR = os.path.join(BASE_DIR, 'output/dengue')


EXPECTED_TEMP_DIR = os.path.abspath(os.path.join(BASE_DIR, 'temp'))
os.makedirs(EXPECTED_TEMP_DIR, exist_ok=True)

# --- Helpers ---
def normalize_key(key: str) -> str:
    return re.sub(r"\s+", " ", key.strip().lower().replace(" de ", " de ").replace(" do ", " do "))

def assert_sheet_exists(workbook, sheet_name):
    assert sheet_name in workbook.sheetnames, f"Sheet '{sheet_name}' not found in workbook"

# --- Fixtures ---
@pytest.fixture
def excel_file():
    pattern = os.path.join(TEMP_DIR, "Epidemiology_Dengue_*.xlsx")
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not files:
        pytest.skip(f"No Excel files matching pattern in {TEMP_DIR}")
    return files[0]

@pytest.fixture
def workbook(excel_file):
    return load_workbook(excel_file)

# --- Tests ---
@pytest.mark.dependency()
def test_update_city_cases(excel_file, workbook):
    assert_sheet_exists(workbook, "City Cases")
    timestamp = datetime.now().strftime("%Y_%m_%d")
    csv_filename = f"city_cases_{timestamp}.csv"

    try:
        csv_path = find_latest_file_in_subfolders(csv_filename, output_dir=OUTPUT_DIR)
    except FileNotFoundError:
        pytest.fail(f"CSV file not found: {csv_filename}")

    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    df.columns = ['Cod Mun', 'Nome Mun', 'Casos Provaveis']
    df = df[df['Cod Mun'].astype(str).str.match(r'^\d{6}$') & (df['Cod Mun'] != '000000')]

    ws = workbook['City Cases']
    excel_data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3)]
    excel_df = pd.DataFrame(excel_data, columns=['Cod Mun', 'Nome Mun', 'Casos Provaveis']).dropna()

    assert len(excel_df) == len(df)

    # Normaliza os tipos da coluna "Casos Provaveis" para int (elimina diferenças entre int e float)
    excel_df['Casos Provaveis'] = pd.to_numeric(excel_df['Casos Provaveis'], errors='coerce').fillna(0).astype(int)
    df['Casos Provaveis'] = pd.to_numeric(df['Casos Provaveis'], errors='coerce').fillna(0).astype(int)

    # Garante que 'Cod Mun' esteja como string em ambos
    excel_df['Cod Mun'] = excel_df['Cod Mun'].astype(str)
    df['Cod Mun'] = df['Cod Mun'].astype(str)

    # Também garante 'Nome Mun' string
    excel_df['Nome Mun'] = excel_df['Nome Mun'].astype(str)
    df['Nome Mun'] = df['Nome Mun'].astype(str)

    # Reset index para evitar problemas na comparação
    excel_df = excel_df.reset_index(drop=True)
    df = df.reset_index(drop=True)

    pd.testing.assert_frame_equal(excel_df, df, check_dtype=False)


@pytest.mark.dependency()
def test_big_numbers_uf(excel_file, workbook):
    assert_sheet_exists(workbook, "Big Numbers UF")
    yaml_path = find_latest_file_in_subfolders('big_numbers_uf_*.yaml', output_dir=OUTPUT_DIR)
    with open(yaml_path, "r", encoding="utf-8") as f:
        yaml_data = yaml.safe_load(f)

    normalized_data = {normalize_key(k): v for k, v in yaml_data.items()}
    ws = workbook["Big Numbers UF"]
    col_map = {'casos prov\u00e1veis de dengue': 'G', 'coeficiente de incid\u00eancia': 'K', '\u00f3bitos em investiga\u00e7\u00e3o': 'P', '\u00f3bitos por dengue': 'Q'}

    for row in ws.iter_rows(min_row=2):
        estado = normalize_key(row[1].value or '')
        if estado not in normalized_data:
            continue
        for key, col in col_map.items():
            if key in normalized_data[estado]:
                expected = normalized_data[estado][key]
                if col == 'G': expected = int(str(expected).replace('.', '').replace(',', ''))
                else:
                    try: expected = float(str(expected).replace(',', '.'))
                    except ValueError: continue
                actual = ws[f"{col}{row[0].row}"].value
                assert actual == expected, f"Mismatch {estado} {key}: {expected} vs {actual}"

@pytest.mark.dependency()
def test_se_completa(excel_file, workbook):
    assert_sheet_exists(workbook, "SE Completa")
    yaml_path = find_latest_file_in_subfolders('semana_epidemiologica_*.yaml', output_dir=OUTPUT_DIR)
    df_yaml = pd.read_csv(yaml_path, dtype=str)
    df_yaml['Ano/Semana'] = df_yaml['Ano/Semana'].str.replace('\xa0', '').str.strip()
    df_yaml["Casos prov\u00e1veis de Dengue"] = pd.to_numeric(df_yaml["Casos prov\u00e1veis de Dengue"], errors="coerce")

    ws = workbook["SE Completa"]
    header_row = next(r for r in ws.iter_rows(min_row=1, max_row=10) if any(c.value == "Ano/Semana" for c in r)).__iter__().__next__().row
    ano_semana_col = [c.column for c in ws[header_row] if c.value == "Ano/Semana"][0]

    start_value = df_yaml['Ano/Semana'].iloc[0]
    start_row = next(cell.row for row in ws.iter_rows(min_row=header_row+1) for cell in row if cell.value == start_value)

    col_map = {cell.value: cell.column for cell in ws[header_row] if cell.value in df_yaml.columns}
    for i, row in df_yaml.iterrows():
        for col_name, col in col_map.items():
            val = int(row[col_name]) if col_name == "Casos prov\u00e1veis de Dengue" else row[col_name]
            excel_val = ws.cell(start_row + i, col).value
            assert excel_val == val

@pytest.mark.dependency()
def test_informe_semana_epidemiologica(excel_file, workbook):
    assert_sheet_exists(workbook, "Informe Semana Epidemiologica")
    yaml_path = find_latest_file_in_subfolders('SE-Y-*.yaml', output_dir=OUTPUT_DIR)
    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)

    se = data['Last_Epidemiological_Week']
    metrics = data['All_Semanas_Data']

    casos = int(metrics['Casos prov\u00e1veis de Dengue'].replace(',', ''))
    incidencia = metrics['Coeficiente de incid\u00eancia'].replace('.', ',').rstrip('-')
    ob_invest = int(metrics['\u00d3bitos em investiga\u00e7\u00e3o'].replace(',', ''))
    ob = int(metrics['\u00d3bitos por Dengue'].replace(',', ''))

    ws = workbook['Informe Semana Epidemiologica']
    rows_with_se = [r for r in range(2, ws.max_row + 1) if str(ws.cell(row=r, column=1).value).startswith("SE ")]
    if not rows_with_se:
        pytest.fail("Nenhuma linha com 'SE ' encontrada na planilha.")

    row = rows_with_se[-1]  # pega a última linha

    assert ws.cell(row=row, column=1).value == f"SE {se}"


    assert ws.cell(row=row, column=1).value == f"SE {se}"
    assert ws.cell(row=row, column=3).value == incidencia
    assert ws.cell(row=row, column=4).value == casos
    assert ws.cell(row=row, column=7).value == ob_invest
    assert ws.cell(row=row, column=8).value == ob

@pytest.mark.dependency()
def test_dengue_cases_year(excel_file, workbook):
    assert_sheet_exists(workbook, "Dengue Cases Year")
    yaml_path = find_latest_file_in_subfolders('big_numbers_uf_*.yaml', output_dir=OUTPUT_DIR)
    with open(yaml_path, "r", encoding="utf-8") as f:
        yaml_data = yaml.safe_load(f)

    total = 0
    for estado, dados in yaml_data.items():
        for k, v in dados.items():
            if normalize_key(k) == "casos prov\u00e1veis de dengue":
                total += int(str(v).replace('.', '').replace(',', ''))

    ws = workbook["Dengue Cases Year"]
    month = datetime.now().strftime("%B").lower()
    row_map = {'january': 2, 'february': 3, 'march': 4, 'april': 5, 'may': 6, 'june': 7,
               'july': 8, 'august': 9, 'september': 10, 'october': 11, 'november': 12, 'december': 13}
    col = next(c for c in range(2, ws.max_column + 1) if str(ws.cell(row=1, column=c).value) == str(datetime.now().year))
    assert ws.cell(row=row_map[month], column=col).value == total

@pytest.mark.dependency(depends=[
    "test_update_city_cases",
    "test_big_numbers_uf",
    "test_se_completa",
    "test_informe_semana_epidemiologica",
    "test_dengue_cases_year"
])

def test_move_file(excel_file):
    results_dir = os.path.join(BASE_DIR, '_results')
    os.makedirs(results_dir, exist_ok=True)
    final_path = os.path.join(results_dir, os.path.basename(excel_file))

    # Move o arquivo para a pasta final
    shutil.move(excel_file, final_path)
    assert os.path.exists(final_path)

    # Se chegou aqui, todos os testes anteriores passaram e o arquivo foi movido com sucesso
    # Agora podemos limpar o diretório TEMP
    for filename in os.listdir(TEMP_DIR):
        file_path = os.path.join(TEMP_DIR, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            pytest.fail(f"Falha ao limpar TEMP_DIR: {e}")