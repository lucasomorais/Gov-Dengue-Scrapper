from openpyxl import load_workbook
from datetime import datetime
import os

def main():
    # Determine current year, month, and date for file name
    current_year = datetime.now().year  # 2025
    current_month = datetime.now().strftime("%B")  # e.g., "April"
    current_date = datetime.now().strftime("%m-%d-%y")  # e.g., "04-11-25"

    # Define the path to the Big_Numbers_UF Excel file
    big_numbers_path = f"_results/Big_Numbers_DATA/Big_Numbers_UF_{current_date}.xlsx"
    
    # Load the Big_Numbers_UF Excel file and sum column A
    try:
        wb_big = load_workbook(big_numbers_path)
        sheet_big = wb_big.active
         
        # Soma ajustada para lidar com vírgulas como separadores decimais
        total_casos = sum(
            float(str(cell.value).replace(',', '.')) if cell.value is not None and str(cell.value).strip().replace('.', '').replace(',', '').isdigit()
            else (cell.value if isinstance(cell.value, (int, float)) else 0)
            for cell in sheet_big["A"][1:]
        )
        
        print(f"Soma final: {total_casos}")
        
    except FileNotFoundError:
        print(f"[ERROR] File not found: {big_numbers_path}")
        return
    except Exception as e:
        print(f"[ERROR] Failed to process {big_numbers_path}: {e}")
        return

    # Load the Dengue_Cases_Year_Copy Excel file as a template
    template_path = "Dengue_Cases_Year/Dengue_Cases_Year_Copy.xlsx"
    wb = load_workbook(template_path)
    sheet = wb.active

    # Find the column index for the current year (2025)
    year_col = None
    for col_idx, cell in enumerate(sheet[1], start=0):  # Row 1 is headers
        if str(cell.value) == str(current_year):
            year_col = col_idx
            break
    if year_col is None:
        print(f"[ERROR] Year {current_year} not found in {template_path}")
        return

    # Find and update the row for the current month
    for row in sheet.iter_rows(min_row=2, values_only=False):  # Skip header
        month = row[0].value  # Column A
        if month == current_month:
            row[year_col].value = total_casos  # Update the 2025 column
            break
    else:
        print(f"[WARNING] Row for {current_month} not found in {template_path}")
        return

    # Define the output path and ensure the directory exists
    output_dir = "_results/Dengue_Cases_Year_DATA"
    os.makedirs(output_dir, exist_ok=True)
    output_path = f"{output_dir}/Dengue_Cases_Year_{current_date}.xlsx"

    # Save the updated Excel file to the new location
    wb.save(output_path)
    print(f"[INFO] Created {output_path} for {current_month} {current_year} with total cases: {total_casos}")

if __name__ == "__main__":
    main()