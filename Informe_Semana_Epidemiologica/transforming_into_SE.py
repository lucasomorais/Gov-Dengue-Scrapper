import yaml
import pandas as pd  # Added for counting rows in CSV
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date, datetime

# Function to count rows in City_Cases CSV
def count_municipios_casos():
    """Count the number of rows in the City_Cases CSV file."""
    current_date = datetime.now().strftime("%m-%d-%y")
    city_cases_path = f"City_Cases-2024/output/City_Cases_{current_date}.csv"
    
    try:
        # Read the CSV file
        df = pd.read_csv(city_cases_path)
        # Return the number of rows (excluding header)
        return len(df)
    except FileNotFoundError:
        print(f"Error: City_Cases CSV file not found at {city_cases_path}")
        exit(1)
    except Exception as e:
        print(f"Error reading City_Cases CSV file: {e}")
        exit(1)

# Step 0: Update the YAML file to use comma as decimal separator for incidencia
# Load the original YAML
with open('Informe_Semana_Epidemiologica/output/SE-Y.yaml', 'r', encoding='utf-8') as f:
    data = yaml.safe_load(f)

# Modify the incidencia value in All_Semanas
for key in data['All_Semanas']:
    if "Coeficiente de incidência" in key:
        value = data['All_Semanas'][key]
        # Ensure the value uses a comma as the decimal separator
        if '.' in value:
            data['All_Semanas'][key] = value.replace('.', ',')  # e.g., '300.8' to '300,8'
        break

# Save the modified YAML
with open('Informe_Semana_Epidemiologica/output/SE-Y.yaml', 'w', encoding='utf-8') as f:
    yaml.dump(data, f, allow_unicode=True, default_flow_style=False)

print("Updated YAML")

# Step 1: Load the YAML file with explicit UTF-8 encoding
yaml_path = 'Informe_Semana_Epidemiologica/output/SE-Y.yaml'
try:
    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)
except FileNotFoundError:
    print(f"Error: YAML file not found at {yaml_path}")
    exit(1)

# Step 2: Extract the epidemiological week and metrics
se_number = data.get('Last_Epidemiological_Week', None)
all_semanas = data.get('All_Semanas', {})

if not se_number or not all_semanas:
    print("Error: Missing 'Last_Epidemiological_Week' or 'All_Semanas' in YAML")
    print(f"YAML content: {data}")
    exit(1)

# Debugging: Print all keys in All_Semanas
print("Keys in All_Semanas:")
for key in all_semanas.keys():
    print(f"- '{key}': '{all_semanas[key]}'")

# Initialize variables to store the metric values
casos_provaveis_str = None
incidencia_str = None
obitos_investigacao_str = None
obitos_str = None

# Define expected metric names
required_metrics = {
    "Casos prováveis de Dengue": "casos_provaveis_str",
    "Coeficiente de incidência": "incidencia_str",
    "Óbitos em investigação - DENV": "obitos_investigacao_str",
    "Óbitos por Dengue": "obitos_str"
}

# Find the keys that contain the desired metric names
for key in all_semanas:
    key_normalized = key.encode().decode('utf-8', errors='replace')  # Handle encoding issues
    for metric_name, var_name in required_metrics.items():
        if metric_name in key_normalized:
            globals()[var_name] = all_semanas[key]
            break

# Check for missing metrics and report them
missing_metrics = [name for name, var in required_metrics.items() if globals()[var] is None]
if missing_metrics:
    print(f"Error: The following required metrics were not found in YAML:")
    for metric in missing_metrics:
        print(f"- {metric}")
    raise ValueError("One or more required metrics not found in YAML data")

# Convert the strings to appropriate types
try:
    casos_provaveis = int(casos_provaveis_str.replace(',', ''))  # e.g., '639,353' to 639353
    # Ensure incidencia_display is a string with comma as decimal separator
    incidencia_display = incidencia_str  # Already '300,8' from YAML
    if '.' in incidencia_display:
        incidencia_display = incidencia_display.replace('.', ',')  # Ensure it's '300,8'
    # Remove trailing "-" if present
    if incidencia_display.endswith('-'):
        incidencia_display = incidencia_display[:-1]  # Remove the last character (e.g., "312,1-" to "312,1")
    obitos_investigacao = int(obitos_investigacao_str)  # e.g., '656' to 656
    obitos = int(obitos_str)  # e.g., '322' to 322
except ValueError as e:
    print(f"Error converting metric values: {e}")
    print(f"casos_provaveis_str: {casos_provaveis_str}")
    print(f"incidencia_str: {incidencia_str}")
    print(f"obitos_investigacao_str: {obitos_investigacao_str}")
    print(f"obitos_str: {obitos_str}")
    exit(1)

# Step 3: Load the Excel workbook and select the sheet
excel_path = 'Informe_Semana_Epidemiologica/copy/Epidemiology - Dengue.xlsx'
try:
    wb = load_workbook(excel_path)
    ws = wb['Informe Semana Epidemiologica']  # Correct sheet name
except FileNotFoundError:
    print(f"Error: Excel file not found at {excel_path}")
    exit(1)
except KeyError:
    print(f"Error: Sheet 'Informe Semana Epidemiologica' not found in the Excel file")
    exit(1)

# Step 4: Find the next available row
last_row = ws.max_row
while last_row > 1 and ws.cell(row=last_row, column=1).value is None:
    last_row -= 1
next_row = last_row + 1

# Step 5: Write data to the respective columns
ws.cell(row=next_row, column=1, value=f"SE {se_number}")  # INFORME: e.g., "SE 11"
ws.cell(row=next_row, column=2, value=date.today().strftime('%d-%b-%Y'))  # DATA: e.g., "19-Mar-2025"

# Column C: INCIDENCIA (write as a string to force display with comma)
ws.cell(row=next_row, column=3, value=incidencia_display)

# Column D: CASOS PROVÁVEIS (formatted as number without decimal places)
ws.cell(row=next_row, column=4, value=casos_provaveis).number_format = '#,##0'

# Column E: MUNICIPIOS CASOS (insert the count from City_Cases)
municipios_casos = count_municipios_casos()  # Call the function to get the count
ws.cell(row=next_row, column=5, value=municipios_casos).number_format = '#,##0'

# Column F: CASOS GRAVES (always 0)
ws.cell(row=next_row, column=6, value=0)

# Column G: OBITOS INVESTIGAÇÃO (formatted as number without decimal places)
ws.cell(row=next_row, column=7, value=obitos_investigacao).number_format = '#,##0'

# Column H: OBITOS (formatted as number without decimal places)
ws.cell(row=next_row, column=8, value=obitos).number_format = '#,##0'

# Column I: LETALIDADE (formatted as percentage with 2 decimal places)
ws.cell(row=next_row, column=9, value=f"=H{next_row}/D{next_row}").number_format = '0.00%'

# Column J: VARIAÇÃO % (always 0%, formatted as percentage without decimal places)
ws.cell(row=next_row, column=10, value=0).number_format = '0%'

# Step 6: Apply center alignment to all cells in the new row
center_alignment = Alignment(horizontal='center')
for col in range(1, 11):
    ws.cell(row=next_row, column=col).alignment = center_alignment

# Step 7: Save the updated workbook to the results folder with the new name
current_date = datetime.now().strftime("%m-%d-%y")
output_excel_path = f'_results/Informe_Semana_Epidemiologica_DATA/Informe_Semana_Epidemiologica_{current_date}.xlsx'
wb.save(output_excel_path)  # Save to results folder with new name
print(f"Excel file updated and saved to {output_excel_path}")