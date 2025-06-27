import yaml, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime


def city_cases():
    municipios = pd.read_csv("output/city_cases_2025_06_27.csv")
    numero_de_municipios = municipios["Cod Mun"].count()
    return (numero_de_municipios)
    

# Load the original YAML
with open('output/SE-Y-25_2025_06_27.yaml', 'r', encoding='utf-8') as f:
    data = yaml.safe_load(f)

# Modify the incidencia value in All_Semanas
for key in data['All_Semanas_Data']:
    if "Coeficiente de incidência" in key:
        value = data['All_Semanas_Data'][key]
        # Ensure the value uses a comma as the decimal separator
        if '.' in value:
            data['All_Semanas_Data'][key] = value.replace('.', ',')  # e.g., '300.8' to '300,8'
        break

# Save the modified YAML
with open('output/SE-Y-25_2025_06_27.yaml', 'w', encoding='utf-8') as f:
    yaml.dump(data, f, allow_unicode=True, default_flow_style=False)

print("Updated YAML")

# Step 1: Load the YAML file with explicit UTF-8 encoding
yaml_path = 'output/SE-Y-25_2025_06_27.yaml'
try:
    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)
except FileNotFoundError:
    print(f"Error: YAML file not found at {yaml_path}")
    exit(1)

# Step 2: Extract the epidemiological week and metrics
se_number = data.get('Last_Epidemiological_Week', None)
all_semanas_data = data.get('All_Semanas_Data', {})

if not se_number or not all_semanas_data:
    print("Error: Missing 'Last_Epidemiological_Week' or 'All_Semanas_Data' in YAML")
    print(f"YAML content: {data}")
    exit(1)

# Debugging: Print all keys in All_Semanas
print("Keys in All_Semanas:")
for key in all_semanas_data.keys():
    print(f"- '{key}': '{all_semanas_data[key]}'")

# Initialize variables to store the metric values
casos_provaveis_str = None
incidencia_str = None
obitos_investigacao_str = None
obitos_str = None

# Define expected metric names
required_metrics = {
    "Casos prováveis de Dengue": "casos_provaveis_str",
    "Coeficiente de incidência": "incidencia_str",
    "Óbitos em investigação": "obitos_investigacao_str",
    "Óbitos por Dengue": "obitos_str"
}

# Find the keys that contain the desired metric names
for key in all_semanas_data:
    key_normalized = key.encode().decode('utf-8', errors='replace')  # Handle encoding issues
    for metric_name, var_name in required_metrics.items():
        if metric_name in key_normalized:
            globals()[var_name] = all_semanas_data[key]
            break

# Check for missing metrics and report them
missing_metrics = [name for name, var in required_metrics.items() if globals()[var] is None]
if missing_metrics:
    print(f"Error: The following required metrics were not found in YAML:")
    for metric in missing_metrics:
        print(f"- {metric}")
    raise ValueError("One or more required metrics not found in YAML data")

# Convert the strings to appropriate types
try:
    casos_provaveis = int(casos_provaveis_str.replace(',', ''))  # Remove vírgulas
    incidencia_display = incidencia_str
    if '.' in incidencia_display:
        incidencia_display = incidencia_display.replace('.', ',')
    if incidencia_display.endswith('-'):
        incidencia_display = incidencia_display[:-1]

    obitos_investigacao = int(obitos_investigacao_str.replace(',', ''))  # Precaução
    obitos = int(obitos_str.replace(',', ''))  # Remove vírgulas
except ValueError as e:
    print(f"Error converting metric values: {e}")
    print(f"casos_provaveis_str: {casos_provaveis_str}")
    print(f"incidencia_str: {incidencia_str}")
    print(f"obitos_investigacao_str: {obitos_investigacao_str}")
    print(f"obitos_str: {obitos_str}")
    exit(1)

# Step 3: Load the Excel workbook and select the sheet
excel_path = '_results/Epidemiology_Dengue_2025_06_20.xlsx'
try:
    wb = load_workbook(excel_path)
    ws = wb['Informe Semana Epidemiologica']  # Correct sheet name
except FileNotFoundError:
    print(f"Error: Excel file not found at {excel_path}")
    exit(1)
except KeyError:
    print(f"Error: Sheet 'Informe Semana Epidemiologica' not found in the Excel file")
    exit(1)

# Step 4: Find the last row with the most recent DATA for valid INFORME
valid_rows = []
for row in range(2, ws.max_row + 1):
    informe_value = ws.cell(row=row, column=1).value
    if informe_value and re.match(r"SE \d+", str(informe_value)):
        data_value = ws.cell(row=row, column=2).value
        if isinstance(data_value, str):
            try:
                data_value = datetime.strptime(data_value, "%d-%b-%Y")
            except ValueError:
                continue
        if isinstance(data_value, datetime):
            valid_rows.append((row, data_value))


# Find the row with the latest date
last_row = max(valid_rows, key=lambda x: x[1])[0]
next_row = last_row + 1
print(f"[INFO] Selected last row: {last_row} with INFORME: {ws.cell(row=last_row, column=1).value} and DATA: {ws.cell(row=last_row, column=2).value}")
print(f"[INFO] Writing new row at: {next_row}")

# Step 5: Write data to the respective columns
ws.cell(row=next_row, column=1, value=f"SE {se_number}")  # INFORME: e.g., "SE 11"
ws.cell(row=next_row, column=2, value=datetime.now()).number_format = 'dd-mmm-yyyy'  # DATA: current date

# Column C: INCIDENCIA (write as a string to force display with comma)
ws.cell(row=next_row, column=3, value=incidencia_display)

# Column D: CASOS PROVÁVEIS (formatted as number without decimal places)
ws.cell(row=next_row, column=4, value=casos_provaveis).number_format = '#,##0'

# Column E: MUNICIPIOS CASOS
ws.cell(row=next_row, column=5, value=city_cases()).number_format = '#,##0'

# Column F: CASOS GRAVES (always 0)
ws.cell(row=next_row, column=6, value=0)

# Column G: OBITOS INVESTIGAÇÃO (formatted as number without decimal places)
ws.cell(row=next_row, column=7, value=obitos_investigacao)

# Column H: OBITOS (formatted as number without decimal places)
ws.cell(row=next_row, column=8, value=obitos)

# Column I: LETALIDADE (formatted as percentage with 2 decimal places)
ws.cell(row=next_row, column=9, value=f"=H{next_row}/D{next_row}").number_format = '0.00%'

# Column J: VARIAÇÃO % (always 0%, formatted as percentage without decimal places)
ws.cell(row=next_row, column=10, value=0).number_format = '0%'

# Step 6: Apply center alignment to all cells in the new row
center_alignment = Alignment(horizontal='center')
for col in range(1, 11):
    ws.cell(row=next_row, column=col).alignment = center_alignment

# Step 7: Clear rows below next_row
if next_row + 1 <= ws.max_row:
    ws.delete_rows(next_row + 1, ws.max_row - next_row)
    print(f"[INFO] Cleared rows from {next_row + 1} to {ws.max_row}")

# Step 8: Save the updated workbook to the results folder with the new name
current_date = datetime.now().strftime("%m_%d_%y")
output_excel_path = f'_results/Informe_Semana_Epidemiologica_{current_date}.xlsx'
wb.save(output_excel_path)  # Save to results folder with new name
print(f"Excel file updated and saved to {output_excel_path}")