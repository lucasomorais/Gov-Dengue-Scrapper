import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from datetime import date

# Load the original YAML
with open('Semana_Epidemiologica/output/SE-Y.yaml', 'r', encoding='utf-8') as f:
    data = yaml.safe_load(f)

# Modify the incidencia value in All_Semanas
for key in data['All_Semanas']:
    if "Coeficiente de incidência" in key:
        value = data['All_Semanas'][key]
        data['All_Semanas'][key] = value.replace('.', ',')  # e.g., '300.8' to '300,8'
        break

# Save the modified YAML
with open('Semana_Epidemiologica/output/SE-Y.yaml', 'w', encoding='utf-8') as f:
    yaml.dump(data, f, allow_unicode=True, default_flow_style=False)


print("updated yaml")

# Step 1: Load the YAML file with explicit UTF-8 encoding
yaml_path = 'Semana_Epidemiologica/output/SE-Y.yaml'
try:
    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)
except FileNotFoundError:
    print(f"Error: YAML file not found at {yaml_path}")
    exit(1)

# Step 2: Extract the epidemiological week and metrics
se_number = data.get('Last_Epidemiological_Week', None)
all_semanas = data.get('All_Semanas', {})

if not se_number or not all_semanas:
    print("Error: Missing 'Last_Epidemiological_Week' or 'All_Semanas' in YAML")
    print(f"YAML content: {data}")
    exit(1)

# Debugging: Print all keys in All_Semanas
print("Keys in All_Semanas:")
for key in all_semanas.keys():
    print(f"- '{key}': '{all_semanas[key]}'")

# Initialize variables to store the metric values
casos_provaveis_str = None
incidencia_str = None
obitos_investigacao_str = None
obitos_str = None

# Define expected metric names
required_metrics = {
    "Casos prováveis de Dengue": "casos_provaveis_str",
    "Coeficiente de incidência": "incidencia_str",
    "Óbitos em investigação - DENV": "obitos_investigacao_str",
    "Óbitos por Dengue": "obitos_str"
}

# Find the keys that contain the desired metric names
for key in all_semanas:
    key_normalized = key.encode().decode('utf-8', errors='replace')  # Handle encoding issues
    for metric_name, var_name in required_metrics.items():
        if metric_name in key_normalized:
            globals()[var_name] = all_semanas[key]
            break

# Check for missing metrics and report them
missing_metrics = [name for name, var in required_metrics.items() if globals()[var] is None]
if missing_metrics:
    print(f"Error: The following required metrics were not found in YAML:")
    for metric in missing_metrics:
        print(f"- {metric}")
    raise ValueError("One or more required metrics not found in YAML data")

# Convert the strings to appropriate types
try:
    casos_provaveis = int(casos_provaveis_str.replace(',', ''))  # e.g., '639,353' to 639353
    incidencia = float(incidencia_str)  # e.g., '300.8' to 300.8 (no need to replace comma here)
    obitos_investigacao = int(obitos_investigacao_str)  # e.g., '656' to 656
    obitos = int(obitos_str)  # e.g., '322' to 322
except ValueError as e:
    print(f"Error converting metric values: {e}")
    print(f"casos_provaveis_str: {casos_provaveis_str}")
    print(f"incidencia_str: {incidencia_str}")
    print(f"obitos_investigacao_str: {obitos_investigacao_str}")
    print(f"obitos_str: {obitos_str}")
    exit(1)

# Step 3: Load the Excel workbook and select the sheet
excel_path = 'Semana_Epidemiologica/copy-to-test/Epidemiology - Dengue.xlsx'
try:
    wb = load_workbook(excel_path)
    ws = wb['Informe Semana Epidemiologica']  # Correct sheet name
except FileNotFoundError:
    print(f"Error: Excel file not found at {excel_path}")
    exit(1)
except KeyError:
    print(f"Error: Sheet 'Informe Semana Epidemiologica' not found in the Excel file")
    exit(1)

# Step 4: Find the next available row
last_row = ws.max_row
while last_row > 1 and ws.cell(row=last_row, column=1).value is None:
    last_row -= 1
next_row = last_row + 1

# Step 5: Write data to the respective columns
ws.cell(row=next_row, column=1, value=f"SE {se_number}")  # INFORME: e.g., "SE 11"
ws.cell(row=next_row, column=2, value=date.today().strftime('%d-%b-%Y'))  # DATA: e.g., "19-Mar-2025"

# Column C: INCIDENCIA (formatted with comma as decimal separator)
ws.cell(row=next_row, column=3, value=incidencia).number_format = '#,##0,0'  # Displays as "300,8"

# Column D: CASOS PROVÁVEIS (formatted as number without decimal places)
ws.cell(row=next_row, column=4, value=casos_provaveis).number_format = '#,##0'

# Column E: MUNICIPIOS CASOS (left empty)
ws.cell(row=next_row, column=5, value="")

# Column F: CASOS GRAVES (always 0)
ws.cell(row=next_row, column=6, value=0)

# Column G: OBITOS INVESTIGAÇÃO (formatted as number without decimal places)
ws.cell(row=next_row, column=7, value=obitos_investigacao).number_format = '#,##0'

# Column H: OBITOS (formatted as number without decimal places)
ws.cell(row=next_row, column=8, value=obitos).number_format = '#,##0'

# Column I: LETALIDADE (formatted as percentage with 2 decimal places)
ws.cell(row=next_row, column=9, value=f"=H{next_row}/D{next_row}").number_format = '0.00%'

# Column J: VARIAÇÃO % (always 0%, formatted as percentage without decimal places)
ws.cell(row=next_row, column=10, value=0).number_format = '0%'

# Step 6: Apply center alignment to all cells in the new row
center_alignment = Alignment(horizontal='center')
for col in range(1, 11):
    ws.cell(row=next_row, column=col).alignment = center_alignment

# Step 7: Save the updated workbook
output_excel_path = 'Semana_Epidemiologica/copy-to-test/Epidemiology - Dengue_updated.xlsx'
wb.save(output_excel_path)
print(f"Excel file successfully updated and saved to {output_excel_path}")