import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# Load the CSV file into a pandas DataFrame
csv_file = "SE_COMPLETA_2023-24/output/SE_COMPLETA_2023-24.csv"
try:
    # Read the CSV file while preserving formatting
    df = pd.read_csv(csv_file, dtype=str, encoding="utf-8")
    
    # Ensure column names are stripped of extra spaces
    df.columns = df.columns.str.strip()
    
    # Strip spaces and unwanted quotes from all string values
    df = df.map(lambda x: x.strip().replace('"', '') if isinstance(x, str) else x)
    
    # Convert 'Ano/Semana' to date format if the column exists
    if 'ANO/SEMANA' in df.columns:
        df['ANO/SEMANA'] = pd.to_datetime(df['ANO/SEMANA'] + '-1', format='%Y/%W-%w', errors='coerce')
    
    # Convert third column to numeric if it exists
    if df.shape[1] >= 3:
        third_column = df.columns[2]
        df[third_column] = pd.to_numeric(df[third_column], errors='coerce')
    
    # Drop empty rows
    df.dropna(inplace=True)
    
    # Calculate the current date
    current_date = datetime.now().strftime("%m-%d-%y")
    
    # Save DataFrame to Excel with dynamic filename
    output_file = f"_results/SE_COMPLETA_2023-24_DATA/SE_COMPLETA_2023-24_{current_date}.xlsx"  # Fixed to use f-string
    df.to_excel(output_file, index=False, engine="openpyxl")
    
    # Load workbook to modify alignment
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Align all cells to the center
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(output_file)
    
    print(f"Successfully converted '{csv_file}' to '{output_file}' while maintaining original properties and formatting.")

except FileNotFoundError:
    print(f"Error: The file '{csv_file}' was not found. Please ensure it exists in the current directory.")
except Exception as e:
    print(f"An error occurred: {e}")