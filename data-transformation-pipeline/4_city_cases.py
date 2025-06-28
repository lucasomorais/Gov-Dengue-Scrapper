import os, sys
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import Alignment


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from modules.utils import (
    get_latest_epidemiology_file,
    find_latest_file_in_subfolders,
    OUTPUT_DIR,
    SOURCE_DIR
)

def update_city_cases():
    try:
        # === [1] Define paths ===
        timestamp = datetime.now().strftime("%Y_%m_%d")
        csv_filename = f"city_cases_{timestamp}.csv"
        csv_path = find_latest_file_in_subfolders(csv_filename, output_dir=OUTPUT_DIR)


        try:
            excel_path = get_latest_epidemiology_file(SOURCE_DIR)
        except FileNotFoundError:
            print(f"[ERROR] No Excel file found in: {SOURCE_DIR}")
            return

        # === [2] Read and clean CSV ===
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        df.columns = df.columns.str.strip().str.lower()

        rename_map = {
            df.columns[0]: "Cod Mun",
            df.columns[1]: "Nome Mun",
            df.columns[2]: "Casos Provaveis"
        }
        df = df.rename(columns=rename_map)
        df = df[df["Cod Mun"].astype(str).str.match(r'^\d{6}$') & (df["Cod Mun"] != "000000")]

        # Optionally rewrite cleaned CSV
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')

        # === [3] Load Excel and get sheet ===
        wb = load_workbook(excel_path)
        try:
            ws = wb["City Cases"]
        except KeyError:
            print(f"[ERROR] Sheet 'City Cases' not found in {excel_path}")
            return

        # === [4] Clear previous rows and inject new data ===
        ws.delete_rows(2, ws.max_row - 1)

        for i, row in df.iterrows():
            excel_row = i + 2
            ws.cell(row=excel_row, column=1, value=row["Cod Mun"])
            ws.cell(row=excel_row, column=2, value=row["Nome Mun"])
            ws.cell(row=excel_row, column=3, value=row["Casos Provaveis"])

        # === [5] Center alignment ===
        align_center = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = align_center

        # === [6] Save and log ===
        wb.save(excel_path)
        print(f"[SUCCESS] Injected {len(df)} rows into 'City Cases' in '{os.path.basename(excel_path)}'.")

    except Exception as e:
        print(f"[ERROR] Failed to clean and inject city case data: {e}")
        raise

if __name__ == "__main__":
    update_city_cases()
