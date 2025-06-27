import pandas as pd
from openpyxl import load_workbook

# File paths
yaml_file = "output/semana_epidemiologica_2025_06_27.yaml"  # Path to the YAML/CSV file
excel_file = "_results/Informe_Semana_Epidemiologica_06_27_25.xlsx"  # Path to the Excel file

# Load YAML data (CSV format assumed from your example)
df_yaml = pd.read_csv(yaml_file, dtype=str)
# Clean columns and values
df_yaml.columns = df_yaml.columns.str.strip()
df_yaml['Ano/Semana'] = df_yaml['Ano/Semana'].str.strip().str.replace('\xa0', '')

# Convert 'Casos prováveis de Dengue' column to numeric (int), coercing errors to NaN
df_yaml["Casos prováveis de Dengue"] = pd.to_numeric(df_yaml["Casos prováveis de Dengue"], errors="coerce")

wb = load_workbook(excel_file)
ws = wb["SE Completa"]

# Find header row and 'Ano/Semana' column index
header_row = None
ano_semana_col = None

for row in ws.iter_rows(min_row=1, max_row=10):  # assuming header in first 10 rows
    for cell in row:
        if cell.value and str(cell.value).strip().lower() == "ano/semana":
            header_row = cell.row
            ano_semana_col = cell.column  # numeric index
            break
    if header_row:
        break

if not ano_semana_col:
    raise Exception("Column 'Ano/Semana' not found in Excel sheet.")

# Normalize the first Ano/Semana in YAML to match
start_ano_semana = df_yaml['Ano/Semana'].iloc[0]

# Find the Excel row where 'Ano/Semana' matches start_ano_semana
start_row = None
for row in ws.iter_rows(min_row=header_row + 1, max_col=ano_semana_col):
    cell = row[ano_semana_col - 1]  # openpyxl columns start at 1
    if cell.value and str(cell.value).strip() == start_ano_semana:
        start_row = cell.row
        break

if not start_row:
    raise Exception(f"Starting Ano/Semana {start_ano_semana} not found in Excel.")

# Map other columns, e.g., 'UF' and 'Casos prováveis de Dengue'
col_map = {}
for cell in ws[header_row]:
    header = str(cell.value).strip() if cell.value else None
    if header in df_yaml.columns:
        col_map[header] = cell.column

# Write YAML data rows into Excel starting at start_row
for i, yaml_row in df_yaml.iterrows():
    excel_row = start_row + i
    for col_name, col_idx in col_map.items():
        val = yaml_row[col_name]
        # Convert 'Casos prováveis de Dengue' values to int if not NaN
        if col_name == "Casos prováveis de Dengue":
            if pd.isna(val):
                val = None
            else:
                val = int(val)
        ws.cell(row=excel_row, column=col_idx, value=val)

# Save the workbook
wb.save(excel_file)
print(f"Updated Excel sheet '{ws.title}' starting at row {start_row} for Ano/Semana {start_ano_semana}.")
