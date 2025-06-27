import yaml
import openpyxl

def normalize_key(key):
    return (
        key.replace(" Do ", " do ")
           .replace(" De ", " de ")
           .strip()
           .lower()
    )

def main():
    # Arquivos
    yaml_file = "output/big_numbers_uf_2025_06_27.yaml"
    excel_file = "_results/Informe_Semana_Epidemiologica_06_27_25.xlsx"
    sheet_name = "Big Numbers UF"

    # Carrega o YAML
    with open(yaml_file, "r", encoding="utf-8") as f:
        yaml_data = yaml.safe_load(f)

    # Normaliza as chaves do YAML
    yaml_data_normalized = {
        normalize_key(k): v for k, v in yaml_data.items()
    }

    # Carrega a planilha
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]

    # Mapas de colunas
    col_map = {
        'casos prováveis de dengue': 'G',
        'coeficiente de incidência': 'K',
        'óbitos em investigação': 'P',
        'óbitos por dengue': 'Q',
    }

    # Itera sobre as linhas da planilha (a partir da 2ª linha)
    for row in ws.iter_rows(min_row=2):
        estado_nome = row[1].value.strip() if row[1].value else None  # Coluna B
        estado_normalizado = normalize_key(estado_nome) if estado_nome else None

        if not estado_normalizado or estado_normalizado not in yaml_data_normalized:
            continue

        dados_estado = yaml_data_normalized[estado_normalizado]
        for campo_yaml, valor in dados_estado.items():
            campo_normalizado = normalize_key(campo_yaml)
            for key_map, col_letter in col_map.items():
                if key_map in campo_normalizado:
                    # Tratamento especial para coluna G (remover separadores de milhar)
                    if col_letter == 'G':
                        print(f"{estado_nome} | original: {valor} -> limpo: {str(valor).replace('.', '').replace(',', '')}")
                        valor_limpo = int(str(valor).replace('.', '').replace(',', ''))
                        ws[f"{col_letter}{row[0].row}"] = valor_limpo
                    else:
                        valor_convertido = float(str(valor).replace(',', '.'))
                        ws[f"{col_letter}{row[0].row}"] = valor_convertido

    # Salva o arquivo modificado
    wb.save(excel_file)
    print(f"[SUCCESS] Planilha atualizada: {excel_file}")

if __name__ == "__main__":
    main()
