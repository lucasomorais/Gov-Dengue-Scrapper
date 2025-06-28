import os, sys, re, yaml
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from modules.utils import (
    find_latest_file_in_subfolders,
    get_latest_epidemiology_file,
    OUTPUT_DIR,
    DENGUE_DIR,
    SOURCE_DIR
)

today_str = datetime.now().strftime("%Y_%m_%d")
file_name = f"city_cases_{today_str}.csv"

def city_cases():

    municipios = pd.read_csv(OUTPUT_DIR + DENGUE_DIR + file_name)
    numero_de_municipios = municipios["Cod Mun"].count()
    return numero_de_municipios

def informe_semana_epidemiologica():
    # Step 1: Find the latest SE-Y-*.yaml file
    try:
        yaml_path = find_latest_file_in_subfolders('SE-Y-*.yaml', output_dir=OUTPUT_DIR)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return

    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
    except FileNotFoundError:
        print(f"Error: YAML file not found at {yaml_path}")
        return

    # Normalize 'Coeficiente de incidência'
    for key in data.get('All_Semanas_Data', {}):
        if "Coeficiente de incidência" in key:
            value = data['All_Semanas_Data'][key]
            if '.' in value:
                data['All_Semanas_Data'][key] = value.replace('.', ',')
            break

    # Save YAML
    with open(yaml_path, 'w', encoding='utf-8') as f:
        yaml.dump(data, f, allow_unicode=True, default_flow_style=False)
    print("Updated YAML")

    # Step 2: Extract values
    se_number = data.get('Last_Epidemiological_Week')
    all_semanas_data = data.get('All_Semanas_Data', {})

    if not se_number or not all_semanas_data:
        print("Error: Missing key data in YAML.")
        return

    print("Keys in All_Semanas:")
    for key, value in all_semanas_data.items():
        print(f"- '{key}': '{value}'")

    # Metrics parsing
    required_metrics = {
        "Casos prováveis de Dengue": None,
        "Coeficiente de incidência": None,
        "Óbitos em investigação": None,
        "Óbitos por Dengue": None
    }

    for key, value in all_semanas_data.items():
        key_normalized = key.encode().decode('utf-8', errors='replace')
        for metric in required_metrics:
            if metric in key_normalized:
                required_metrics[metric] = value

    if any(v is None for v in required_metrics.values()):
        print("Missing required metrics in YAML:")
        for k, v in required_metrics.items():
            if v is None:
                print(f"- {k}")
        return

    try:
        casos_provaveis = int(required_metrics["Casos prováveis de Dengue"].replace(',', ''))
        incidencia_display = required_metrics["Coeficiente de incidência"].replace('.', ',').rstrip('-')
        obitos_investigacao = int(required_metrics["Óbitos em investigação"].replace(',', ''))
        obitos = int(required_metrics["Óbitos por Dengue"].replace(',', ''))
    except ValueError as e:
        print(f"Error converting values: {e}")
        return

    # Step 3: Load Excel
    try:
        excel_path = get_latest_epidemiology_file('_results')
        wb = load_workbook(excel_path)
        ws = wb['Informe Semana Epidemiologica']
    except (FileNotFoundError, KeyError) as e:
        print(f"Error loading workbook or sheet: {e}")
        return

    # Step 4: Find last valid row
    valid_rows = []
    for row in range(2, ws.max_row + 1):
        informe = ws.cell(row=row, column=1).value
        data_val = ws.cell(row=row, column=2).value
        if informe and re.match(r"SE \d+", str(informe)):
            if isinstance(data_val, str):
                try:
                    data_val = datetime.strptime(data_val, "%d-%b-%Y")
                except ValueError:
                    continue
            if isinstance(data_val, datetime):
                valid_rows.append((row, data_val))

    if not valid_rows:
        print("Error: No valid rows found in Excel.")
        return

    last_row = max(valid_rows, key=lambda x: x[1])[0]
    next_row = last_row + 1
    print(f"[INFO] Selected last row: {last_row} → Writing to row: {next_row}")

    # Step 5: Write new row
    ws.cell(row=next_row, column=1, value=f"SE {se_number}")
    ws.cell(row=next_row, column=2, value=datetime.now()).number_format = 'dd-mmm-yyyy'
    ws.cell(row=next_row, column=3, value=incidencia_display)
    ws.cell(row=next_row, column=4, value=casos_provaveis).number_format = '#,##0'
    ws.cell(row=next_row, column=5, value=city_cases()).number_format = '#,##0'
    ws.cell(row=next_row, column=6, value=0)
    ws.cell(row=next_row, column=7, value=obitos_investigacao)
    ws.cell(row=next_row, column=8, value=obitos)
    ws.cell(row=next_row, column=9, value=f"=H{next_row}/D{next_row}").number_format = '0.00%'
    ws.cell(row=next_row, column=10, value=0).number_format = '0%'

    # Step 6: Center alignment
    center = Alignment(horizontal='center')
    for col in range(1, 11):
        ws.cell(row=next_row, column=col).alignment = center

    # Step 7: Clear excess rows
    if next_row + 1 <= ws.max_row:
        ws.delete_rows(next_row + 1, ws.max_row - next_row)
        print(f"[INFO] Cleared rows from {next_row + 1} to {ws.max_row}")

    # Step 8: Save
    today_str = datetime.now().strftime("%Y_%m_%d")
    output_path = os.path.join(SOURCE_DIR, f"Epidemiology_Dengue_{today_str}.xlsx")
    wb.save(output_path)
    print(f"✅ Excel file updated and saved to {output_path}")


if __name__ == "__main__":
    informe_semana_epidemiologica()
