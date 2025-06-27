import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def main():
    try:
        # Correct way to get current script's directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = script_dir  # You are running from dash_gerencial directly

        # Paths and timestamp
        timestamp = datetime.now().strftime("%Y_%m_%d")
        csv_path = os.path.join(project_root, "output", f"city_cases_{timestamp}.csv")
        excel_path = os.path.join(project_root, "_results", "Informe_Semana_Epidemiologica_06_27_25.xlsx")

        # Read and clean CSV
        df = pd.read_csv(csv_path, encoding='utf-8-sig')
        df.columns = df.columns.str.strip().str.lower()

        rename_map = {
            df.columns[0]: "Cod Mun",
            df.columns[1]: "Nome Mun",
            df.columns[2]: "Casos Provaveis"
        }
        df = df.rename(columns=rename_map)

        df = df[df["Cod Mun"].astype(str).str.match(r'^\d{6}$') & (df["Cod Mun"] != "000000")]

        # Optionally rewrite cleaned CSV
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')

        # Load Excel and get "City Cases" sheet
        wb = load_workbook(excel_path)
        ws = wb["City Cases"]

        # Clear previous data rows
        ws.delete_rows(2, ws.max_row - 1)

        # Inject rows
        for i, row in df.iterrows():
            excel_row = i + 2
            ws.cell(row=excel_row, column=1, value=row["Cod Mun"])
            ws.cell(row=excel_row, column=2, value=row["Nome Mun"])
            ws.cell(row=excel_row, column=3, value=row["Casos Provaveis"])

        # Align to center
        align_center = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = align_center

        wb.save(excel_path)
        print(f"[SUCCESS] Injected {len(df)} rows into 'City Cases' in '{os.path.basename(excel_path)}'.")

    except Exception as e:
        print(f"[ERROR] Failed to clean and inject city case data: {e}")
        raise

if __name__ == "__main__":
    main()
