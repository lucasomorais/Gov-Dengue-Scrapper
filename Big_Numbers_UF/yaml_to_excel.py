import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def yaml_to_excel_with_exact_formatting(yaml_file, excel_file):
    """
    Converts a YAML file containing UF data into an Excel file with the exact format
    matching Pasta1.xlsx (blue headers, alternating row colors, centered data, gridlines).
    Uses "Geral" format with commas as decimal separators.
    """
    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"

    # Define styles for headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define border style for gridlines
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # Add headers
    headers = ["Casos 2025", "Incidence 2025", "Óbitos 2025 (em investigação)", "Óbitos 2025 (confirmados)"]
    ws.append(headers)

    # Format header row
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    ws.row_dimensions[1].height = 22  # Match header height from Pasta1.xlsx

    # Load YAML data
    with open(yaml_file, "r", encoding="utf-8") as f:
        uf_data = yaml.safe_load(f)

    # Define alternating row colors
    light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Light blue
    light_orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light orange
    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows with alternating colors
    for row_idx, (uf, data) in enumerate(uf_data.items(), 2):
        casos_raw = data.get("Casos prováveis de Dengue", "")
        incidencia_raw = data.get("Coeficiente de incidência", "")
        obitos_investigacao_raw = data.get("Óbitos em investigação", "")
        obitos_confirmados_raw = data.get("Óbitos por Dengue", "")

        # Sanitize and convert
        casos = casos_raw.replace(",", "") if casos_raw else "0"
        incidencia = incidencia_raw.replace(".", ",") if incidencia_raw else "0"
        obitos_investigacao = obitos_investigacao_raw.replace(",", "") if obitos_investigacao_raw else "0"
        obitos_confirmados = obitos_confirmados_raw.replace(",", "") if obitos_confirmados_raw else "0"

        # Prepare row data
        row = [
            str(int(casos)) if casos else "0",
            incidencia,
            str(int(obitos_investigacao)) if obitos_investigacao else "0",
            str(int(obitos_confirmados)) if obitos_confirmados else "0"
        ]
        ws.append(row)

        # Apply alternating row colors and formatting
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.number_format = "General"
            if row_idx % 2 == 0:
                cell.fill = light_blue_fill
            else:
                cell.fill = light_orange_fill
        ws.row_dimensions[row_idx].height = 16

    # Adjust column widths dynamically
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 15)
        ws.column_dimensions[column].width = adjusted_width

    # Enable gridlines
    ws.sheet_view.showGridLines = True

    # Save the workbook
    wb.save(excel_file)
    print(f"Excel file saved to {excel_file}")

# Calculate the current date and use it in the filename
current_date = datetime.now().strftime("%m-%d-%y")
yaml_to_excel_with_exact_formatting(
    "Big_Numbers_UF/output/dengue_uf_data.yaml", 
    f"_results/Big_Numbers_DATA/Big_Numbers_UF_{current_date}.xlsx"
)
