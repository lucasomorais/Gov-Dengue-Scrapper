import os, sys
import pandas as pd
from openpyxl import load_workbook


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from modules.utils import (
    find_latest_file_in_subfolders,
    get_latest_epidemiology_file,
    OUTPUT_DIR,
    TEMP_DIR
)

def se_completa():
    # === [1] Encontrar o arquivo CSV mais recente (semana_epidemiologica_*.yaml) ===
    try:
        yaml_path = find_latest_file_in_subfolders('semana_epidemiologica_*.yaml', output_dir=OUTPUT_DIR)
    except FileNotFoundError as e:
        print(f"[ERROR] YAML/CSV file not found: {e}")
        return

    # === [2] Carregar os dados CSV ===
    df_yaml = pd.read_csv(yaml_path, dtype=str)
    df_yaml.columns = df_yaml.columns.str.strip()
    df_yaml['Ano/Semana'] = df_yaml['Ano/Semana'].str.strip().str.replace('\xa0', '', regex=False)
    df_yaml["Casos prováveis de Dengue"] = pd.to_numeric(df_yaml["Casos prováveis de Dengue"], errors="coerce")

    # === [3] Encontrar o Excel mais recente ===
    try:
        excel_path = get_latest_epidemiology_file(TEMP_DIR)
        wb = load_workbook(excel_path)
        ws = wb["SE Completa"]
    except FileNotFoundError:
        print(f"[ERROR] Excel file not found in: {TEMP_DIR}")
        return
    except KeyError:
        print(f"[ERROR] Sheet 'SE Completa' not found in {excel_path}")
        return

    # === [4] Identificar a linha de cabeçalho e coluna 'Ano/Semana' ===
    header_row = None
    ano_semana_col = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == "ano/semana":
                header_row = cell.row
                ano_semana_col = cell.column
                break
        if header_row:
            break

    if not ano_semana_col:
        raise Exception("Column 'Ano/Semana' not found in Excel sheet.")

    start_ano_semana = df_yaml['Ano/Semana'].iloc[0]

    # === [5] Encontrar a linha no Excel correspondente ao primeiro 'Ano/Semana' ===
    start_row = None
    for row in ws.iter_rows(min_row=header_row + 1, max_col=ano_semana_col):
        cell = row[ano_semana_col - 1]
        if cell.value and str(cell.value).strip() == start_ano_semana:
            start_row = cell.row
            break

    if not start_row:
        raise Exception(f"Starting Ano/Semana '{start_ano_semana}' not found in Excel.")

    # === [6] Mapear colunas do Excel que batem com o CSV ===
    col_map = {}
    for cell in ws[header_row]:
        header = str(cell.value).strip() if cell.value else None
        if header in df_yaml.columns:
            col_map[header] = cell.column

    # === [7] Escrever dados na planilha
    for i, yaml_row in df_yaml.iterrows():
        excel_row = start_row + i
        for col_name, col_idx in col_map.items():
            val = yaml_row[col_name]
            if col_name == "Casos prováveis de Dengue":
                val = int(val) if pd.notna(val) else None
            ws.cell(row=excel_row, column=col_idx, value=val)

    # === [8] Salvar planilha ===
    wb.save(excel_path)
    print(f"✅ Planilha '{ws.title}' atualizada a partir da linha {start_row} (Ano/Semana: {start_ano_semana})")

if __name__ == "__main__":
    se_completa()
