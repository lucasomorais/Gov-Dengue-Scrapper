import os, yaml, sys
from datetime import datetime
from openpyxl import load_workbook
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

from modules.utils import (
    get_latest_epidemiology_file, OUTPUT_DIR
)


def normalize_key(key):
    return key.split(":")[0].strip().lower()

def parse_int(value):
    if value is None:
        return 0
    try:
        return int(str(value).replace('.', '').replace(',', ''))
    except ValueError:
        return 0

def dengue_cases_year():
    # === [1] Locate the most recent YAML file ===
    try:
        yaml_path = next(
            f for f in sorted(
                (os.path.join(root, file)
                 for root, _, files in os.walk(OUTPUT_DIR)
                 for file in files if file.startswith("big_numbers_uf_") and file.endswith(".yaml")),
                reverse=True
            )
        )
    except StopIteration:
        print("[ERROR] No YAML file found.")
        return

    with open(yaml_path, "r", encoding="utf-8") as f:
        yaml_data = yaml.safe_load(f)

    total_casos = 0
    for estado, dados in yaml_data.items():
        if isinstance(dados, dict):
            for campo_yaml, valor in dados.items():
                if normalize_key(campo_yaml) == "casos prováveis de dengue":
                    total_casos += parse_int(valor)
                    break

    # === [2] Locate the latest Excel epidemiology file ===
    try:
        excel_path = get_latest_epidemiology_file("_results")
    except FileNotFoundError:
        print("[ERROR] No Excel file found in _results")
        return

    wb = load_workbook(excel_path)
    try:
        ws = wb["Dengue Cases Year"]
    except KeyError:
        print(f"[ERROR] Sheet 'Dengue Cases Year' not found in {excel_path}")
        return

    # === [3] Map month to row (January = 2, ..., December = 13) ===
    month_row_map = {
        'january': 2, 'february': 3, 'march': 4, 'april': 5,
        'may': 6, 'june': 7, 'july': 8, 'august': 9,
        'september': 10, 'october': 11, 'november': 12, 'december': 13
    }
    current_year = datetime.now().year
    current_month = datetime.now().strftime("%B").lower()

    # === [4] Find column for current year ===
    year_col = None
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).strip() == str(current_year):
            year_col = col
            break

    if not year_col:
        print(f"[ERROR] Year column {current_year} not found in sheet header.")
        return

    # === [5] Write the total into the correct cell ===
    row = month_row_map.get(current_month)
    if not row:
        print(f"[ERROR] Unknown month: {current_month}")
        return

    ws.cell(row=row, column=year_col, value=total_casos)
    wb.save(excel_path)
    print(f"[SUCCESS] Inserted {total_casos} into 'Dengue Cases Year' -> row {row}, col {year_col} ({current_month.capitalize()} {current_year})")


if __name__ == "__main__":
    dengue_cases_year()
